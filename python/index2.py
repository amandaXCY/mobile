from bs4 import BeautifulSoup
import os
import xlwt
import re
from openpyxl import Workbook
import pandas as pd
from python.variable import save_path, xlsx_path, xlsx_title


class ParseHTML:
    def __init__(self):

        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = '数据产品经理'
        self.title = xlsx_title
        self.coloumIndex = 1;
        self.write_xlw(self.title, 1)
        self.open_file()

        path = save_path
        print(111)

        if not os.path.exists(path):  # 判断当前路径是否存在，没有则创建new文件夹
            os.makedirs(path)

        self.workbook.save(xlsx_path)

      
        #self.workbook.save('./dist/file/job-back.xlsx')

    def open_file(self):
        path = './python/html'
        files = os.listdir(path);
        index = 0;
        for file in files:
            if file != ".DS_Store":
                if index > 0:
                    self.coloumIndex = self.coloumIndex + 29

                index = index + 1
                file = open(path + "/" + file)
                str_html = file.read()
                
                soup = BeautifulSoup(str_html, 'lxml')  # lxml为解析器
                self.parse_list2(soup)

    def parse_list2(self, soup):

        self.parse_job(soup)
        self.parse_company(soup)
        self.parse_scale(soup)
        self.parse_trades(soup)
        self.parse_salary(soup)
        self.parse_ee(soup)
        self.parse_tags(soup)
        self.parse_welfare(soup)
        self.parse_joburl(soup)

    # title
    def write_xlw(self, title, col):
       
        for row in range(0, len(title)):
            self.worksheet.cell(row=col, column=row+1, value="{0}".format(title[row]))

    def write_xlw_row(self, title, row ,color ='black'):
        for i in range(len(title)):
            print(title[i])
            self.worksheet.cell(row = i + self.coloumIndex + 1,column= row+1, value="{0}".format(title[i]))

    # 职位名称
    def parse_job(self, soup):

        tags = []
        nodes = soup.find_all('span', {"class": 'job-name'})
        for node in nodes:
            tags.append(node.a.attrs['title'])

        self.write_xlw_row(tags, self.title.index('职位名称'))

    # 公司名称
    def parse_company(self,soup):

        tags = []
        nodes = soup.find_all('div', {"class":'company-text'})
        for node in nodes:
            tags.append(node.a.attrs['title'])
        self.write_xlw_row(tags, self.title.index('公司名称'))

    # 规模
    def parse_scale(self,soup):
        tags = []
        nodes = soup.find_all('div', {"class": 'company-text'})

        for node in nodes:
            pattern = re.compile(r'(\d+(.*))')  # 用于匹配至少一个数字

            m = re.split(pattern,node.get_text())
            tags.append(m[1])
        self.write_xlw_row(tags, self.title.index('规模'))

    # 行业
    def parse_trades(self,soup):
        tags = []
        nodes = soup.find_all('a', {"class": 'false-link'})
        for node in nodes:
            tags.append(node.get_text())
        self.write_xlw_row(tags, self.title.index('行业'))

    # 薪资
    def parse_salary(self,soup):
        tags = []
        year_salary = []
        nodes = soup.select('.job-limit > span.red')

        for node in nodes:

            text = node.get_text().split('·')
            tags.append(text[0])

            if len(text) == 2:
                year_salary.append(text[1])
            else:
                year_salary.append("")

        self.write_xlw_row(tags, self.title.index('薪资'))
        self.write_xlw_row(year_salary, self.title.index('年薪'))

    # 经验和学习Experience、Education
    def parse_ee(self,soup):
        experience = []
        education = []
        nodes = soup.select('.job-limit p')
        for node in nodes:
            contents = node.contents;

            contents_len = len(contents)
            if contents_len == 3:
                experience.append(contents[0]);
                education.append(contents[2]);

            elif contents_len == 2:
                if '年' in contents[0]:
                    experience.append(contents[0])
                    education.append("");
                elif '经验不限' in contents[0]:
                    experience.append(contents[0])
                    education.append("");
                else:
                    experience.append("");
                    education.append(contents[0])

        self.write_xlw_row(experience, self.title.index('经验'))
        self.write_xlw_row(education, self.title.index('学历'))

    def parse_tags(self, soup):

        tag_nodes = soup.select('.tags')

        tags = []
        for node in tag_nodes:
            tags.append(node.get_text()[1:].replace('\n','、'))

        self.write_xlw_row(tags, self.title.index('标签'))

    #job detali
    def parse_joburl(self,soup):
        nodes = soup.select('.job-name a')
        tags = []
        for node in nodes:
            link = node.attrs['href'];
            name ="详情"
            path = 'https://www.zhipin.com'+link
            #url = HYPERLINK("%s","%s")' % ('https://www.zhipin.com'+link, name)
            url = '=HYPERLINK("{url}","{name}")'.format(url=path,name=name)

            tags.append(url)

        self.write_xlw_row(tags, self.title.index('job地址'), 'blue')

    # 福利
    def parse_welfare(self,soup):
        tag_nodes = soup.select('.info-append .info-desc')
        tags = []
        for node in tag_nodes:
            tags.append(node.get_text())
        self.write_xlw_row(tags, self.title.index('福利'))

    # 福利
    def parse_welfare(self, soup):
        tag_nodes = soup.select('.info-append .info-desc')
        tags = []
        for node in tag_nodes:
            tags.append(node.get_text())

        self.write_xlw_row(tags, self.title.index('福利'))

